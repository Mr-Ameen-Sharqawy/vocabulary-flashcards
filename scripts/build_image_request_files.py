#!/usr/bin/env python3
"""Build hand-off files for generating the remaining Primary 4 vocabulary images."""

from __future__ import annotations

import csv
import re
from pathlib import Path
from textwrap import dedent

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
MANIFEST = ROOT / "cartoon_image_manifest.txt"
OUT_XLSX = ROOT / "Primary4_image_requests_remaining.xlsx"
OUT_CSV = ROOT / "Primary4_image_requests_remaining.csv"
OUT_MD = ROOT / "Primary4_image_requests_remaining_batches.md"

BASE_STYLE = (
    "Clean 2D educational cartoon illustration for a Primary 4 English vocabulary flashcard; "
    "portrait 3:4 composition; warm pale-cream background; one focused child-friendly subject or scene; "
    "rounded friendly shapes; clean dark navy outlines; soft gentle shadows; bright but balanced colors; "
    "no text, English or Arabic letters, numbers, labels, logos, watermark, border, collage, split panel, or extra objects."
)

# Scene guidance is used for concepts and phrases that are difficult to depict literally.
SCENE_HINTS = {
    "absent": "an empty school desk while the other children are learning in class",
    "acts of kindness": "a child helping an elderly neighbour carry grocery bags",
    "adventure": "two young explorers with a map and a small backpack on a safe nature trail",
    "air pollution": "a child looking at grey smoke above a busy city road, with blue sky farther away",
    "amazing": "a delighted child looking at a bright rainbow after rain",
    "announce": "a friendly school child speaking into a microphone to share news",
    "annoying": "a child gently covering their ears because of a buzzing fly nearby",
    "annual": "a cheerful school celebration that happens once each year, with simple decorations but no words",
    "apologize": "one child kindly saying sorry to another child after a small accident",
    "areas": "a simple community park with clearly separate playground, garden, and picnic areas",
    "as well": "two children both joining the same friendly activity",
    "at night": "a child looking through a window at the moon and stars at night",
    "attention": "a child listening carefully while a teacher points to a picture board with no writing",
    "behavior": "children showing polite classroom behaviour by waiting their turn",
    "between classes": "children walking calmly in a school hallway between two lessons",
    "brave": "a brave child calmly helping a frightened friend during a safe situation",
    "break time": "children enjoying a school break in the playground",
    "bullying": "a kind teacher stopping one child from teasing another child, shown safely and gently",
    "busy market": "a colourful Egyptian market with many shoppers and stalls",
    "by hand": "a child carefully making a craft with their hands",
    "call out": "a child calling loudly to a friend across a safe playground",
    "care": "a child gently caring for a small plant",
    "care for = take care of": "a child feeding and caring for a pet rabbit",
    "carelessly": "a child leaving toys scattered carelessly on the floor while an adult reminds them gently",
    "celebrate": "children happily celebrating together with simple paper decorations",
    "celebrated": "children happily celebrating together with simple paper decorations",
    "celebration": "a happy school celebration with balloons and a small cake but no words",
    "challenges": "a child climbing a small safe practice wall with a teacher encouraging them",
    "character": "a cheerful storybook character with a clear, simple costume",
    "check": "a child carefully checking a school bag before leaving",
    "checked": "a child carefully checking a school bag before leaving",
    "clean": "a child cleaning a classroom table with a cloth",
    "climbed": "a child safely climbing a playground ladder",
    "come inside": "a friendly child opening a house door and welcoming a visitor inside",
    "come out": "a small fox coming out of a cave",
    "competition": "children taking part in a friendly school drawing competition, no writing",
    "covered with": "a toy box clearly covered with a colourful blanket",
    "create": "a child creating a colourful paper craft at a table",
    "created": "a child creating a colourful paper craft at a table",
    "crooked": "a simple crooked wooden shelf beside a straight shelf",
    "culture": "an Egyptian family sharing traditional music, food, and clothing",
    "cycle": "a child riding a bicycle safely with a helmet",
    "cycled": "a child riding a bicycle safely with a helmet",
    "daily routine": "a simple morning routine scene with a child getting dressed and packing a school bag",
    "danger": "a child standing safely far from a clear warning situation such as a hot fire",
    "daytime": "a bright daytime scene with sun in the sky and children outdoors",
    "deep breath": "a child taking a slow deep breath in a calm garden",
    "disappearing": "a small animal slowly disappearing behind a bush",
    "discuss": "two children discussing a picture together at a classroom table",
    "discussed": "two children discussing a picture together at a classroom table",
    "do a great job": "a teacher giving a happy thumbs-up to a child who finished a neat craft",
    "do well": "a proud child completing a school task successfully",
    "during the day": "a bright daytime scene with sun in the sky and children outdoors",
    "eating habits": "a child choosing healthy foods including vegetables and water",
    "endangered": "a protected animal with a forest ranger watching over it",
    "energy": "a healthy child running and playing energetically outside",
    "enjoy": "a smiling child enjoying a book in a cosy reading corner",
    "enjoyed": "a smiling child enjoying a book in a cosy reading corner",
    "enter": "a child entering a classroom through an open door",
    "entered": "a child entering a classroom through an open door",
    "especially": "a child pointing excitedly to their favourite item in a group of toys",
    "everywhere": "colourful autumn leaves scattered everywhere in a park",
    "excellent": "a smiling teacher giving a gold star sticker to a child, no writing",
    "exciting": "children excitedly opening a treasure map for a safe adventure",
    "expensive": "a shiny toy displayed behind a shop window with large coins nearby",
    "experience": "a child trying a new safe activity such as planting a seed for the first time",
    "face": "a child facing a small challenge with a calm confident expression",
    "faded": "two cloth pieces, one bright and one faded in colour",
    "faraway": "a child looking through binoculars toward a faraway mountain",
    "feel": "a child showing a clear happy feeling with a hand over their heart",
    "felt": "a child showing a clear happy feeling with a hand over their heart",
    "fewer": "two simple groups of apples, one group clearly having fewer apples",
    "final": "a child reaching the final step of a simple race",
    "first time": "a child proudly trying a bicycle for the first time with a parent nearby",
    "fit": "a child checking that a helmet fits comfortably",
    "fixed": "a carpenter fixing a small wooden chair with simple tools",
    "fluffy tail": "a fox with a large fluffy tail in a forest",
    "follow": "a child following a parent along a safe path",
    "followed": "a child following a parent along a safe path",
    "for free": "a kind librarian giving a child a book for free, no signs or text",
    "for special events": "a family wearing festive clothes for a special celebration",
    "forever": "two best friends making a promise under a sunny sky",
    "fresh": "fresh vegetables and fruit on a clean kitchen table",
    "gentle": "a child gently holding a small bird in their hands",
    "getting enough sleep": "a child sleeping peacefully in bed with a moon and stars outside",
    "greedy": "a child trying to keep too many sweets while friends wait, shown gently",
    "grew": "a small green plant growing taller from a seed",
    "hard": "a child touching a hard wooden block",
    "healthy": "a healthy child eating vegetables and drinking water",
    "heavy": "a child and an adult lifting one heavy box together safely",
    "helping hands": "several children joining hands to help clean a community park",
    "high": "a kite flying high above a playground",
    "hungry": "a hungry child looking at a healthy sandwich on a plate",
    "impressed by": "a child looking impressed by a tall waterfall",
    "in bright colors": "a row of crafts painted in bright colours",
    "in danger": "a small animal safely protected by a forest ranger from danger",
    "in groups": "children working together in small groups at classroom tables",
    "instead of": "a child choosing fruit instead of candy, shown as two clear choices",
    "invite": "a child inviting a friend to a birthday party with a blank card",
    "invited": "a child inviting a friend to a birthday party with a blank card",
    "It's fun to": "children smiling while doing a fun outdoor activity together",
    "jingling": "small silver bells gently jingling on a colourful ribbon",
    "keep": "a child carefully keeping coins in a small jar",
    "kept": "a child carefully keeping coins in a small jar",
    "knew": "a child recognising the correct answer in a picture quiz with no words",
    "known for": "a famous Egyptian place shown with its distinctive local craft",
    "laugh at": "children kindly laughing together at a funny puppet show, not at a person",
    "light": "a child easily lifting one light feather",
    "lives": "different animal homes in a simple nature scene",
    "local": "local farmers selling fresh food at a small neighbourhood market",
    "look dirty": "a muddy pair of shoes that look dirty beside clean shoes",
    "loud": "a child covering their ears near a loud drum",
    "low": "a low table beside a tall table",
    "make a difference": "children cleaning litter from a park and making it look better",
    "make friends": "two shy children smiling and becoming friends in a playground",
    "marks": "clear animal footprints marking a path in soft sand",
    "mind": "a child thinking carefully with a small thought cloud showing simple shapes, no words",
    "mistakes": "a child gently erasing a small drawing mistake and trying again",
    "move around": "a small monkey moving around tree branches",
    "moved": "a child moving a chair to a new place",
    "natural": "a natural forest scene with trees, rocks, and a stream",
    "near": "two houses standing near each other",
    "needed": "a child needing and receiving a small help from a teacher",
    "notice": "a child noticing a small butterfly on a flower",
    "noticed": "a child noticing a small butterfly on a flower",
    "offer": "a child offering a glass of water to a thirsty friend",
    "offered": "a child offering a glass of water to a thirsty friend",
    "on the road": "a family driving safely on a clear road",
    "one by one": "children standing in a neat line and taking turns one by one",
    "paid for": "a child giving coins to buy bread at a small shop, no signs",
    "painted": "a child painting a colourful flower picture",
    "prepare": "a family preparing a healthy meal together",
    "prepared": "a family preparing a healthy meal together",
    "pretend": "a child pretending to be a doctor with a toy stethoscope",
    "pretended": "a child pretending to be a doctor with a toy stethoscope",
    "proud to": "a child proudly holding a finished craft project",
    "quiet": "a child reading quietly in a library corner",
    "ran": "a child running safely in a playground",
    "reach": "a child reaching a high shelf with a safe small step stool",
    "reached": "a child reaching a high shelf with a safe small step stool",
    "real": "a child comparing a real apple with a toy apple",
    "reason": "a child thinking about a reason for a simple event with a light bulb idea, no words",
    "reduce": "a child reducing waste by using a reusable bottle",
    "reduced": "a child reducing waste by using a reusable bottle",
    "regret": "a child looking sorry after breaking a small toy, with a parent helping",
    "request": "a polite child raising a hand to make a request in class",
    "safe": "a child wearing a helmet and knee pads while cycling safely",
    "safely": "a child crossing a road safely with an adult at a zebra crossing",
    "said": "a child talking kindly to a friend using a speech bubble with no letters",
    "same": "two matching red balls side by side",
    "shake": "a child gently shaking a small bottle of juice",
    "share with": "two children sharing crayons at a table",
    "sharp": "a child looking alert while listening to a teacher",
    "shock": "a surprised child reacting to a safe small shock such as a balloon popping",
    "shook": "a child gently shaking a small bottle of juice",
    "shy": "a shy child standing quietly behind a parent before meeting a friend",
    "sick": "a sick child resting in bed while a caring parent brings water",
    "sick people": "a nurse kindly caring for sick people in a hospital room",
    "simple": "a very simple paper airplane made from one sheet of paper",
    "skilled": "a skilled carpenter carefully making a small wooden chair",
    "slept": "a child sleeping peacefully in bed with a moon outside",
    "smart": "a smart child solving a simple wooden puzzle",
    "soft": "a child touching a soft fluffy pillow",
    "solutions": "children working together to solve a simple classroom problem",
    "solve": "a child solving a simple wooden puzzle",
    "solved": "a child solving a simple wooden puzzle",
    "special": "a child holding a special handmade gift",
    "stay safe": "children following safe rules near a road with an adult",
    "stay sharp": "a child alert and focused while using a magnifying glass to observe a leaf",
    "still stand": "an old stone building that is still standing strongly",
    "strong": "a strong child carrying a light school bag confidently",
    "studied": "a child studying with books and a lamp at a desk",
    "study": "a child studying with books and a lamp at a desk",
    "stylish": "a child wearing a stylish colourful outfit",
    "suddenly": "a child suddenly surprised by a butterfly flying nearby",
    "take a shower": "a child taking a shower, shown respectfully from behind with a towel nearby",
    "take out .... from": "a child taking a book out from a school bag",
    "take the bus": "a child getting on a yellow school bus safely",
    "talk about": "two children talking about a picture book together",
    "taste better": "a child smiling after adding a healthy topping to a sandwich",
    "terrible": "a child looking worried about terrible rain outside a window, shown safely",
    "That's cool!": "two children looking excited at a cool science model, with no words",
    "the king of the jungle land": "a proud lion standing in a sunny grassland",
    "thick": "a thick book beside a thin book",
    "threw": "a child throwing a soft ball safely in a playground",
    "throw": "a child throwing a soft ball safely in a playground",
    "took": "a child taking a book from a shelf",
    "traditional": "an Egyptian child wearing traditional clothing during a cultural celebration",
    "traditions": "an Egyptian family sharing traditional food and music",
    "unkindness": "a child choosing kindness and comforting a sad friend after unkind behaviour",
    "usual": "a child following their usual school morning routine",
    "wake up early": "a child waking up early as morning sunlight enters the bedroom",
    "walk through": "a family walking through a forest path",
    "walked": "a child walking along a safe path",
    "weak": "a small weak plant needing water beside a strong healthy plant",
    "What a beautiful day!": "children happily enjoying a beautiful sunny day in a green park",
    "wild": "wild animals living freely in a forest",
    "winner": "a proud child holding a gold medal after a friendly school competition",
    "wise": "a wise old owl reading a blank book in a tree",
    "wonderful": "a child looking delighted at a wonderful colourful garden",
    "work as": "an adult working as a forest guide, helping children on a safe nature walk",
    "Yours truly": "a child placing a blank friendly letter into an envelope, with no words",
    "yummy": "a smiling child enjoying a yummy bowl of fruit salad",
}


def slugify(term: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", term.lower()).strip("-")
    return slug or "vocabulary"


def parse_manifest() -> list[dict[str, str | int]]:
    rows: list[dict[str, str | int]] = []
    pattern = re.compile(r"^\s*(\d+)\.\s*(.*?)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*$")
    for raw_line in MANIFEST.read_text(encoding="utf-8").splitlines():
        match = pattern.match(raw_line)
        if not match:
            continue
        number, term, arabic, unit_lesson = match.groups()
        index = int(number)
        if index <= 10:
            continue
        rows.append(
            {
                "id": index,
                "term": term,
                "arabic": arabic,
                "unit_lesson": unit_lesson,
                "filename": f"cartoon-{index:03d}-{slugify(term)}.png",
            }
        )
    return rows


def visual_idea(term: str, arabic: str) -> str:
    return SCENE_HINTS.get(
        term,
        f"a single clear, child-friendly visual that shows the literal meaning of “{term}” ({arabic})",
    )


def image_prompt(term: str, arabic: str, idea: str) -> str:
    return (
        f"{BASE_STYLE} Target word or phrase: “{term}”. "
        f"Arabic teaching meaning for interpretation only: “{arabic}”. "
        f"Show: {idea}."
    )


def write_csv(rows: list[dict[str, str | int]]) -> None:
    fields = [
        "image_id",
        "batch",
        "unit_lesson",
        "term_en",
        "translation_ar",
        "suggested_filename",
        "visual_idea_en",
        "generation_prompt_en",
        "status",
    ]
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        for offset, row in enumerate(rows):
            idea = visual_idea(str(row["term"]), str(row["arabic"]))
            writer.writerow(
                {
                    "image_id": row["id"],
                    "batch": offset // 50 + 1,
                    "unit_lesson": row["unit_lesson"],
                    "term_en": row["term"],
                    "translation_ar": row["arabic"],
                    "suggested_filename": row["filename"],
                    "visual_idea_en": idea,
                    "generation_prompt_en": image_prompt(str(row["term"]), str(row["arabic"]), idea),
                    "status": "To generate",
                }
            )


def write_markdown(rows: list[dict[str, str | int]]) -> None:
    content = [
        "# Primary 4 — طلبات صور المفردات المتبقية",
        "",
        "هذا الملف يحتوي على **459** مفردة تحتاج صورًا فردية. استخدم سطر الطلب الجاهز لكل كلمة لإنتاج صورة واحدة، ثم احفظ الصورة باسم الملف المقترح. الصور يجب أن تكون كرتونية وتعليمية ونظيفة **من دون أي كتابة أو شعار أو علامة مائية**.",
        "",
        "> لا تضع أكثر من مفردة في صورة واحدة. أرسل الصور لاحقًا بأسماء الملفات المقترحة، وسأربطها تلقائيًا ببطاقات التطبيق.",
        "",
        "## أسلوب ثابت لكل الصور",
        "",
        f"`{BASE_STYLE}`",
        "",
    ]
    for batch_start in range(0, len(rows), 50):
        batch = rows[batch_start : batch_start + 50]
        batch_number = batch_start // 50 + 1
        content.extend(
            [
                f"## الدفعة {batch_number:02d} — الكلمات {batch[0]['id']:03d} إلى {batch[-1]['id']:03d}",
                "",
                "| # | الكلمة | الترجمة | اسم الملف المقترح | فكرة الصورة |",
                "|---:|---|---|---|---|",
            ]
        )
        for row in batch:
            idea = visual_idea(str(row["term"]), str(row["arabic"]))
            content.append(
                f"| {row['id']:03d} | {row['term']} | {row['arabic']} | `{row['filename']}` | {idea} |"
            )
        content.extend(["", "### النصوص الجاهزة لهذه الدفعة", ""])
        for row in batch:
            idea = visual_idea(str(row["term"]), str(row["arabic"]))
            content.extend(
                [
                    f"#### {row['id']:03d} — {row['term']}",
                    "",
                    f"**اسم الملف:** `{row['filename']}`",
                    "",
                    f"```text\n{image_prompt(str(row['term']), str(row['arabic']), idea)}\n```",
                    "",
                ]
            )
    OUT_MD.write_text("\n".join(content), encoding="utf-8")


def write_workbook(rows: list[dict[str, str | int]]) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "README"
    overview.sheet_view.rightToLeft = True
    overview.append(["ملف طلبات صور مفردات الصف الرابع"])
    overview.append(["عدد الصور المطلوبة", len(rows)])
    overview.append(["حجم الدفعة", "50 صورة"])
    overview.append(["طريقة الاستخدام", "ولّد صورة واحدة لكل صف، واحفظها باسم Suggested filename، ثم ارفع الصور في أرشيف ZIP."])
    overview.append(["شرط أساسي", "لا كتابة، لا شعارات، لا علامة مائية، ولا صور مجمّعة أو مقسمة."])
    overview.append(["ملاحظة", "الترجمة العربية للفهم فقط، والنص الإنجليزي الجاهز هو الذي يُنسخ إلى مولد الصور."])
    overview["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    overview["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    overview.merge_cells("A1:B1")
    overview.column_dimensions["A"].width = 25
    overview.column_dimensions["B"].width = 88
    for cell in overview["A"]:
        cell.font = Font(bold=True, color="1F1F1F")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in overview["B"]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    headers = [
        "رقم الصورة",
        "الدفعة",
        "الوحدة / الدرس",
        "الكلمة الإنجليزية",
        "الترجمة العربية",
        "اسم الملف المقترح",
        "فكرة الصورة (EN)",
        "النص الجاهز للتوليد (EN)",
        "الحالة",
    ]
    widths = [12, 10, 16, 27, 28, 38, 55, 105, 16]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    alternating_fill = PatternFill("solid", fgColor="EAF3F8")
    thin = Side(style="thin", color="C7D6E2")
    border = Border(bottom=thin)

    for batch_start in range(0, len(rows), 50):
        batch = rows[batch_start : batch_start + 50]
        sheet = workbook.create_sheet(f"Batch {batch_start // 50 + 1:02d}")
        sheet.sheet_view.rightToLeft = True
        sheet.freeze_panes = "A2"
        sheet.append(headers)
        for col_index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col_index)].width = width
            cell = sheet.cell(row=1, column=col_index)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for offset, row in enumerate(batch):
            idea = visual_idea(str(row["term"]), str(row["arabic"]))
            sheet.append(
                [
                    row["id"],
                    batch_start // 50 + 1,
                    row["unit_lesson"],
                    row["term"],
                    row["arabic"],
                    row["filename"],
                    idea,
                    image_prompt(str(row["term"]), str(row["arabic"]), idea),
                    "To generate",
                ]
            )
            row_number = offset + 2
            sheet.row_dimensions[row_number].height = 56
            for column in range(1, len(headers) + 1):
                cell = sheet.cell(row=row_number, column=column)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                if offset % 2 == 1:
                    cell.fill = alternating_fill
        sheet.auto_filter.ref = f"A1:I{len(batch)+1}"

    workbook.save(OUT_XLSX)


def main() -> None:
    rows = parse_manifest()
    if len(rows) != 459:
        raise SystemExit(f"Expected 459 remaining rows after excluding 001–010; got {len(rows)}")
    write_csv(rows)
    write_markdown(rows)
    write_workbook(rows)
    print(f"Created {OUT_XLSX.name}, {OUT_CSV.name}, and {OUT_MD.name} with {len(rows)} rows.")


if __name__ == "__main__":
    main()
